"""
backend-tests/conftest.py
pytest hooks for the VitalCore backend test suite.
The pytest_sessionfinish hook MUST live here (not in a test file).
"""
import json
import sys
import pathlib
from datetime import datetime


def pytest_sessionfinish(session, exitstatus):
    """After all tests: dump _results to JSON and generate HTML + Excel reports."""
    # Lazily import openpyxl only when needed
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        HAS_OPENPYXL = True
    except ImportError:
        HAS_OPENPYXL = False

    # Grab results from the test module
    # We import it here so we share the already-populated list from the run
    try:
        import importlib.util, os
        spec = importlib.util.spec_from_file_location(
            "test_backend",
            os.path.join(os.path.dirname(__file__), "test_backend.py")
        )
        mod = importlib.util.module_from_spec(spec)
        # We do NOT exec_module() here – that would re-run tests!
        # Instead we reach into the already-loaded module via sys.modules
        import sys as _sys
        tb = None
        for name, m in _sys.modules.items():
            if hasattr(m, "_results") and hasattr(m, "BASE_URL") and hasattr(m, "record"):
                tb = m
                break
    except Exception:
        tb = None

    results = tb._results if tb is not None else []
    base_url = tb.BASE_URL if tb is not None else "http://127.0.0.1:3000"

    reports_dir = pathlib.Path("backend-tests/reports")
    reports_dir.mkdir(parents=True, exist_ok=True)

    # JSON
    json_path = reports_dir / "backend_results.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2)

    total = len(results)
    passed = sum(1 for r in results if r["Pass/Fail"] == "PASS")
    failed = total - passed
    pct = round(passed / total * 100, 1) if total > 0 else 0

    # HTML
    rows_html = ""
    for r in results:
        color = "#d4edda" if r["Pass/Fail"] == "PASS" else "#f8d7da"
        bg = "#28a745" if r["Pass/Fail"] == "PASS" else "#dc3545"
        badge = (
            f'<span style="background:{bg};color:#fff;'
            f'padding:2px 8px;border-radius:4px;">{r["Pass/Fail"]}</span>'
        )
        rows_html += (
            f'<tr style="background:{color}">'
            f'<td>{r["Test ID"]}</td><td>{r["Test Case"]}</td>'
            f'<td style="font-size:11px">{r["Endpoint"]}</td>'
            f'<td style="font-size:10px">{r["Input"]}</td>'
            f'<td style="font-size:10px">{r["Expected Result"]}</td>'
            f'<td style="font-size:10px">{r["Actual Result"]}</td>'
            f'<td>{r["HTTP Status"]}</td><td>{badge}</td>'
            f'<td style="font-size:10px;color:#c00">{r["Error Details"]}</td>'
            f'<td>{r["Execution Time (s)"]}s</td></tr>'
        )

    html = f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<title>VitalCore Backend Test Report</title>
<style>
body{{font-family:Arial,sans-serif;margin:24px;background:#f5f5f5}}
h1{{color:#1a1a2e}}
table{{border-collapse:collapse;width:100%;background:#fff;font-size:12px}}
th{{background:#1a1a2e;color:#fff;padding:8px;text-align:left}}
td{{border:1px solid #ccc;padding:5px 7px;vertical-align:top}}
.badges{{display:flex;gap:12px;margin:16px 0}}
.badge{{padding:10px 18px;border-radius:8px;color:#fff;font-size:15px;font-weight:bold}}
</style></head><body>
<h1>VitalCore Backend API Test Report</h1>
<p><strong>Generated:</strong> {datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")}</p>
<p><strong>Target:</strong> {base_url}</p>
<div class="badges">
<div class="badge" style="background:#1a1a2e">Total: {total}</div>
<div class="badge" style="background:#28a745">Passed: {passed}</div>
<div class="badge" style="background:#dc3545">Failed: {failed}</div>
<div class="badge" style="background:#007bff">Pass Rate: {pct}%</div>
</div>
<table><thead><tr>
<th>Test ID</th><th>Test Case</th><th>Endpoint</th><th>Input</th>
<th>Expected</th><th>Actual</th><th>HTTP</th><th>Result</th><th>Error</th><th>Time</th>
</tr></thead><tbody>{rows_html}</tbody></table>
</body></html>"""

    html_path = reports_dir / "backend_results.html"
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)

    # Excel
    if HAS_OPENPYXL:
        thin = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill("solid", fgColor="1A1A2E")

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(["VitalCore Backend API Test Summary"])
        ws.append(["Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")])
        ws.append(["Target URL", base_url])
        ws.append([])
        ws.append(["Metric", "Value"])
        ws.append(["Total Test Cases", total])
        ws.append(["Executed", total])
        ws.append(["Passed", passed])
        ws.append(["Failed", failed])
        ws.append(["Pass Rate (%)", pct])
        ws["A1"].font = Font(bold=True, size=14)
        for cell in ws[5]:
            cell.font = hdr_font
            cell.fill = hdr_fill
        for row in ws.iter_rows(min_row=6, max_row=10, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin
        ws["B8"].fill = PatternFill("solid", fgColor="D4EDDA")
        ws["B9"].fill = PatternFill("solid", fgColor="F8D7DA" if failed > 0 else "D4EDDA")
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40

        ws2 = wb.create_sheet("Detailed Results")
        cols = [
            "Test ID", "Test Case", "Endpoint", "Input",
            "Expected Result", "Actual Result", "HTTP Status",
            "Pass/Fail", "Error Details", "Execution Time (s)",
        ]
        ws2.append(cols)
        for cell in ws2[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin
        for rec in results:
            ws2.append([rec.get(c, "") for c in cols])
            rn = ws2.max_row
            rf = PatternFill("solid", fgColor=("D4EDDA" if rec["Pass/Fail"] == "PASS" else "F8D7DA"))
            for ci in range(1, len(cols) + 1):
                cell = ws2.cell(row=rn, column=ci)
                cell.fill = rf
                cell.border = thin
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for i, w in enumerate([10, 45, 32, 22, 30, 30, 12, 10, 30, 14], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        xlsx_path = reports_dir / "backend_results.xlsx"
        wb.save(xlsx_path)
    else:
        xlsx_path = None

    print(f"\n{'='*60}")
    print("  BACKEND TEST SUMMARY")
    print(f"{'='*60}")
    print(f"  Total  : {total}")
    print(f"  Passed : {passed}")
    print(f"  Failed : {failed}")
    print(f"  Rate   : {pct}%")
    print(f"  JSON   : {json_path}")
    print(f"  HTML   : {html_path}")
    if xlsx_path:
        print(f"  Excel  : {xlsx_path}")
    print(f"{'='*60}\n")
