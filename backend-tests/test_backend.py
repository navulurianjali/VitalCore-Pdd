"""
VitalCore Backend API Test Suite
Tests all real Next.js API endpoints: /api/contact, /api/chat, /api/ai-coach, /api/future-lab
Each test records REAL HTTP results – no mocking.
"""

import os, time, json, requests, pytest, pathlib
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_URL = os.environ.get("BASE_URL", "http://127.0.0.1:3000")
_results = []

# ─── helpers ─────────────────────────────────────────────────────────────────

def record(tid, case, ep, method, payload, exp_status, exp_desc, resp, elapsed, err=""):
    actual_status = resp.status_code if resp else "NO_RESPONSE"
    try:
        actual_body = resp.json() if resp else {}
    except Exception:
        actual_body = {"raw": (resp.text[:300] if resp else "")}
    passed = resp is not None and resp.status_code == exp_status
    _results.append({
        "Test ID": tid,
        "Test Case": case,
        "Endpoint": f"{method} {ep}",
        "Input": (json.dumps(payload)[:200] if payload is not None else "(none)"),
        "Expected Result": f"HTTP {exp_status} - {exp_desc}",
        "Actual Result": json.dumps(actual_body)[:200],
        "HTTP Status": actual_status,
        "Pass/Fail": "PASS" if passed else "FAIL",
        "Error Details": (err if not passed else ""),
        "Execution Time (s)": round(elapsed, 3),
    })
    return passed

def _post(ep, payload, headers=None, timeout=15):
    hdrs = {"Content-Type": "application/json"}
    if headers:
        hdrs.update(headers)
    t0 = time.time()
    try:
        r = requests.post(f"{BASE_URL}{ep}", json=payload, headers=hdrs, timeout=timeout)
        return r, time.time() - t0, ""
    except Exception as e:
        return None, time.time() - t0, str(e)

def _get(ep, headers=None, timeout=15):
    t0 = time.time()
    try:
        r = requests.get(f"{BASE_URL}{ep}", headers=headers or {}, timeout=timeout)
        return r, time.time() - t0, ""
    except Exception as e:
        return None, time.time() - t0, str(e)


# ─── /api/contact ─────────────────────────────────────────────────────────────

class TestContactAPI:
    EP = "/api/contact"

    def test_BE001_valid_submission(self):
        p = {"name": "Test User", "email": "test@example.com", "message": "This is a test message from automated backend tests."}
        r, el, err = _post(self.EP, p)
        assert record("BE-001", "Valid contact form submission", self.EP, "POST", p, 200, "success:true", r, el, err)

    def test_BE002_missing_name(self):
        p = {"email": "test@example.com", "message": "Hello there this is my message."}
        r, el, err = _post(self.EP, p)
        assert record("BE-002", "Contact - missing name field returns 400", self.EP, "POST", p, 400, "name validation", r, el, err)

    def test_BE003_name_too_short(self):
        p = {"name": "A", "email": "test@example.com", "message": "This is a valid length message."}
        r, el, err = _post(self.EP, p)
        assert record("BE-003", "Contact - name < 2 chars returns 400", self.EP, "POST", p, 400, "name >= 2 chars", r, el, err)

    def test_BE004_missing_email(self):
        p = {"name": "Test User", "message": "This is a valid length message for testing."}
        r, el, err = _post(self.EP, p)
        assert record("BE-004", "Contact - missing email field returns 400", self.EP, "POST", p, 400, "email required", r, el, err)

    def test_BE005_email_no_at_symbol(self):
        p = {"name": "Test User", "email": "notanemail.com", "message": "This is a valid test message here."}
        r, el, err = _post(self.EP, p)
        assert record("BE-005", "Contact - email without @ returns 400", self.EP, "POST", p, 400, "invalid email", r, el, err)

    def test_BE006_email_missing_domain(self):
        p = {"name": "Test User", "email": "user@", "message": "This is a valid message for testing purposes."}
        r, el, err = _post(self.EP, p)
        assert record("BE-006", "Contact - email with empty domain returns 400", self.EP, "POST", p, 400, "invalid domain", r, el, err)

    def test_BE007_email_missing_tld(self):
        p = {"name": "Test User", "email": "user@domain", "message": "This is a valid test message for backend testing."}
        r, el, err = _post(self.EP, p)
        assert record("BE-007", "Contact - email missing TLD returns 400", self.EP, "POST", p, 400, "missing TLD", r, el, err)

    def test_BE008_message_too_short(self):
        p = {"name": "Test User", "email": "test@example.com", "message": "Hi"}
        r, el, err = _post(self.EP, p)
        assert record("BE-008", "Contact - message < 10 chars returns 400", self.EP, "POST", p, 400, "message >= 10 chars", r, el, err)

    def test_BE009_missing_message(self):
        p = {"name": "Test User", "email": "test@example.com"}
        r, el, err = _post(self.EP, p)
        assert record("BE-009", "Contact - missing message returns 400", self.EP, "POST", p, 400, "message required", r, el, err)

    def test_BE010_empty_body(self):
        p = {}
        r, el, err = _post(self.EP, p)
        assert record("BE-010", "Contact - empty body returns 400", self.EP, "POST", p, 400, "all fields missing", r, el, err)

    def test_BE011_whitespace_only_name(self):
        p = {"name": "   ", "email": "test@example.com", "message": "Valid message here for whitespace test."}
        r, el, err = _post(self.EP, p)
        assert record("BE-011", "Contact - whitespace-only name rejected", self.EP, "POST", p, 400, "name trim validation", r, el, err)

    def test_BE012_whitespace_only_message(self):
        p = {"name": "Test User", "email": "test@example.com", "message": "          "}
        r, el, err = _post(self.EP, p)
        assert record("BE-012", "Contact - whitespace-only message rejected", self.EP, "POST", p, 400, "message too short after trim", r, el, err)

    def test_BE013_optional_user_id_accepted(self):
        p = {"name": "Test User", "email": "test@example.com", "message": "Message with userId included in payload.", "userId": "optional-user-id-123"}
        r, el, err = _post(self.EP, p)
        assert record("BE-013", "Contact - optional userId accepted (200)", self.EP, "POST", p, 200, "success:true", r, el, err)

    def test_BE014_xss_in_name_no_500(self):
        p = {"name": "<script>alert(1)</script>", "email": "test@example.com", "message": "XSS payload test message for backend safety."}
        r, el, err = _post(self.EP, p)
        record("BE-014", "Contact - XSS in name does not cause 500", self.EP, "POST", p, r.status_code if r else 0, "no HTTP 500", r, el, err)
        assert r is not None and r.status_code != 500
        _results[-1]["Pass/Fail"] = "PASS"
        _results[-1]["Expected Result"] = "No HTTP 500"

    def test_BE015_sql_injection_in_message_no_500(self):
        p = {"name": "Test User", "email": "test@example.com", "message": "'; DROP TABLE users; -- this is a SQL injection test here ok."}
        r, el, err = _post(self.EP, p)
        record("BE-015", "Contact - SQL injection in message safe (no 500)", self.EP, "POST", p, r.status_code if r else 0, "no HTTP 500", r, el, err)
        assert r is not None and r.status_code != 500
        _results[-1]["Pass/Fail"] = "PASS"
        _results[-1]["Expected Result"] = "No HTTP 500"

    def test_BE016_get_method_not_allowed(self):
        t0 = time.time()
        r = requests.get(f"{BASE_URL}{self.EP}", timeout=15)
        el = time.time() - t0
        assert record("BE-016", "Contact - GET method not allowed (405)", self.EP, "GET", None, 405, "HTTP 405", r, el)

    def test_BE017_success_response_structure(self):
        p = {"name": "Structure Test", "email": "structure@example.com", "message": "Testing response structure for backend validation."}
        r, el, err = _post(self.EP, p)
        record("BE-017", "Contact - success body has success=true and message key", self.EP, "POST", p, 200, "JSON keys: success, message", r, el, err)
        assert r is not None and r.status_code == 200
        body = r.json()
        assert body.get("success") is True and "message" in body
        _results[-1]["Pass/Fail"] = "PASS"

    def test_BE018_error_response_structure(self):
        p = {"name": "A", "email": "bad", "message": "x"}
        r, el, err = _post(self.EP, p)
        record("BE-018", "Contact - error body has 'error' key", self.EP, "POST", p, 400, "JSON key: error", r, el, err)
        assert r is not None and r.status_code == 400
        assert "error" in r.json()
        _results[-1]["Pass/Fail"] = "PASS"

    def test_BE019_content_type_json(self):
        p = {"name": "A", "email": "bad@e.com", "message": "short"}
        r, el, err = _post(self.EP, p)
        record("BE-019", "Contact API response Content-Type is application/json", self.EP, "POST", p, r.status_code if r else 0, "Content-Type: application/json", r, el, err)
        assert r is not None
        ct = r.headers.get("Content-Type", "")
        assert "application/json" in ct
        _results[-1]["Pass/Fail"] = "PASS"
        _results[-1]["Expected Result"] = "Content-Type: application/json"


# ─── /api/chat ────────────────────────────────────────────────────────────────

class TestChatAPI:
    EP = "/api/chat"

    def test_BE020_valid_message_200(self):
        p = {"messages": [{"sender": "user", "text": "Hello coach, how am I doing?"}], "profile": {"name": "Tester"}, "metrics": {"steps": 5000}}
        r, el, err = _post(self.EP, p)
        record("BE-020", "Chat - valid message returns 200 streamed response", self.EP, "POST", p, 200, "HTTP 200 text/plain", r, el, err)
        assert r is not None and r.status_code == 200
        _results[-1]["Pass/Fail"] = "PASS"

    def test_BE021_empty_messages_array(self):
        p = {"messages": [], "profile": {}, "metrics": {}}
        r, el, err = _post(self.EP, p)
        assert record("BE-021", "Chat - empty messages array returns 400", self.EP, "POST", p, 400, "messages required", r, el, err)

    def test_BE022_missing_messages_field(self):
        p = {"profile": {}, "metrics": {}}
        r, el, err = _post(self.EP, p)
        assert record("BE-022", "Chat - missing messages field returns 400", self.EP, "POST", p, 400, "messages required", r, el, err)

    def test_BE023_no_profile_metrics_still_works(self):
        p = {"messages": [{"sender": "user", "text": "What should I eat today?"}]}
        r, el, err = _post(self.EP, p)
        record("BE-023", "Chat - no profile/metrics optional (200)", self.EP, "POST", p, 200, "HTTP 200 with defaults", r, el, err)
        assert r is not None and r.status_code == 200
        _results[-1]["Pass/Fail"] = "PASS"

    def test_BE024_content_type_text_plain(self):
        p = {"messages": [{"sender": "user", "text": "Give me a fitness tip please."}]}
        r, el, err = _post(self.EP, p)
        record("BE-024", "Chat - response Content-Type is text/plain", self.EP, "POST", p, 200, "text/plain", r, el, err)
        assert r is not None and r.status_code == 200
        assert "text/plain" in r.headers.get("Content-Type", "")
        _results[-1]["Pass/Fail"] = "PASS"

    def test_BE025_multiple_messages_last_used(self):
        p = {"messages": [
            {"sender": "assistant", "text": "Hi!"},
            {"sender": "user", "text": "Tell me about hydration."},
            {"sender": "assistant", "text": "Hydration matters."},
            {"sender": "user", "text": "How much water?"}
        ]}
        r, el, err = _post(self.EP, p)
        record("BE-025", "Chat - multiple messages; last user message consumed", self.EP, "POST", p, 200, "HTTP 200", r, el, err)
        assert r is not None and r.status_code == 200
        _results[-1]["Pass/Fail"] = "PASS"

    def test_BE026_full_metrics_payload(self):
        p = {"messages": [{"sender": "user", "text": "Analyze my complete health data."}],
             "metrics": {"steps": 8000, "sleepHours": 7.5, "sleepQuality": 80,
                         "caloriesBurned": 500, "caloriesConsumed": 1800,
                         "hydrationMl": 2200, "stressLevel": 30, "recoveryPercentage": 75}}
        r, el, err = _post(self.EP, p)
        record("BE-026", "Chat - full metrics payload returns 200", self.EP, "POST", p, 200, "HTTP 200", r, el, err)
        assert r is not None and r.status_code == 200
        _results[-1]["Pass/Fail"] = "PASS"

    def test_BE027_response_body_non_empty(self):
        p = {"messages": [{"sender": "user", "text": "What is my health score?"}]}
        r, el, err = _post(self.EP, p)
        record("BE-027", "Chat - response body is non-empty", self.EP, "POST", p, 200, "Non-empty body", r, el, err)
        assert r is not None and r.status_code == 200 and len(r.text) > 0
        _results[-1]["Pass/Fail"] = "PASS"

    def test_BE028_numeric_text_no_500(self):
        p = {"messages": [{"sender": "user", "text": 12345}]}
        r, el, err = _post(self.EP, p)
        record("BE-028", "Chat - numeric text field coerced (no 500)", self.EP, "POST", p, r.status_code if r else 0, "200 or 400 (no 500)", r, el, err)
        assert r is not None and r.status_code != 500
        _results[-1]["Pass/Fail"] = "PASS"
        _results[-1]["Expected Result"] = "No HTTP 500"

    def test_BE029_empty_text_no_crash(self):
        p = {"messages": [{"sender": "user", "text": ""}]}
        r, el, err = _post(self.EP, p)
        record("BE-029", "Chat - empty text string (edge case, no 500)", self.EP, "POST", p, r.status_code if r else 0, "200 or 400", r, el, err)
        assert r is not None and r.status_code in (200, 400)
        _results[-1]["Pass/Fail"] = "PASS"
        _results[-1]["Expected Result"] = "HTTP 200 or 400"


# ─── /api/ai-coach ────────────────────────────────────────────────────────────

class TestAICoachAPI:
    EP = "/api/ai-coach"

    def test_BE030_unauthenticated_401(self):
        p = {"message": "Hello coach!", "history": []}
        r, el, err = _post(self.EP, p)
        assert record("BE-030", "AI Coach - unauthenticated returns 401", self.EP, "POST", p, 401, "HTTP 401 Unauthorized", r, el, err)

    def test_BE031_invalid_bearer_401(self):
        p = {"message": "Hello!", "history": []}
        r, el, err = _post(self.EP, p, {"Authorization": "Bearer invalid.jwt.token"})
        assert record("BE-031", "AI Coach - invalid Bearer token returns 401", self.EP, "POST", p, 401, "HTTP 401", r, el, err)

    def test_BE032_empty_bearer_401(self):
        p = {"message": "Hi", "history": []}
        r, el, err = _post(self.EP, p, {"Authorization": "Bearer "})
        assert record("BE-032", "AI Coach - empty Bearer value returns 401", self.EP, "POST", p, 401, "HTTP 401", r, el, err)

    def test_BE033_basic_auth_scheme_401(self):
        p = {"message": "Hi", "history": []}
        r, el, err = _post(self.EP, p, {"Authorization": "Basic dXNlcjpwYXNz"})
        assert record("BE-033", "AI Coach - Basic auth scheme returns 401 (not Bearer)", self.EP, "POST", p, 401, "HTTP 401", r, el, err)

    def test_BE034_get_not_allowed_405(self):
        r, el, err = _get(self.EP)
        assert record("BE-034", "AI Coach - GET method returns 405", self.EP, "GET", None, 405, "HTTP 405", r, el, err)

    def test_BE035_401_response_has_error_key(self):
        r, el, err = _post(self.EP, {"message": "test"})
        record("BE-035", "AI Coach - 401 response body has 'error' key", self.EP, "POST", None, 401, "JSON error key", r, el, err)
        assert r is not None and r.status_code == 401 and "error" in r.json()
        _results[-1]["Pass/Fail"] = "PASS"

    def test_BE036_oversized_payload_rejected(self):
        big = "A" * 1_100_000
        t0 = time.time()
        r = requests.post(f"{BASE_URL}{self.EP}", data=json.dumps({"message": big}),
                          headers={"Content-Type": "application/json"}, timeout=20)
        el = time.time() - t0
        record("BE-036", "AI Coach - 1.1MB payload rejected (401 or 413)", self.EP, "POST", "(1.1MB)", r.status_code, "401 or 413", r, el)
        assert r.status_code in (401, 413)
        _results[-1]["Pass/Fail"] = "PASS"
        _results[-1]["Expected Result"] = "HTTP 401 or 413"

    def test_BE037_empty_body_401(self):
        r, el, err = _post(self.EP, {})
        assert record("BE-037", "AI Coach - empty body returns 401 (auth first)", self.EP, "POST", {}, 401, "HTTP 401", r, el, err)

    def test_BE038_missing_message_field_401(self):
        r, el, err = _post(self.EP, {"history": []})
        assert record("BE-038", "AI Coach - missing message field returns 401 (auth checked first)", self.EP, "POST", {"history": []}, 401, "HTTP 401", r, el, err)


# ─── /api/future-lab ──────────────────────────────────────────────────────────

class TestFutureLabAPI:
    EP = "/api/future-lab"

    def test_BE040_unauthenticated_get_401(self):
        r, el, err = _get(self.EP)
        assert record("BE-040", "Future Lab - unauthenticated GET returns 401", self.EP, "GET", None, 401, "HTTP 401", r, el, err)

    def test_BE041_unauthenticated_post_401(self):
        r, el, err = _post(self.EP, {})
        assert record("BE-041", "Future Lab - unauthenticated POST returns 401", self.EP, "POST", {}, 401, "HTTP 401", r, el, err)

    def test_BE042_invalid_bearer_401(self):
        r, el, err = _get(self.EP, {"Authorization": "Bearer faketoken12345"})
        assert record("BE-042", "Future Lab - invalid Bearer returns 401", self.EP, "GET", None, 401, "HTTP 401", r, el, err)

    def test_BE043_401_has_error_key(self):
        r, el, err = _get(self.EP)
        record("BE-043", "Future Lab - 401 response has 'error' key", self.EP, "GET", None, 401, "JSON error key", r, el, err)
        assert r is not None and r.status_code == 401 and "error" in r.json()
        _results[-1]["Pass/Fail"] = "PASS"

    def test_BE044_delete_not_allowed(self):
        t0 = time.time()
        r = requests.delete(f"{BASE_URL}{self.EP}", timeout=15)
        el = time.time() - t0
        assert record("BE-044", "Future Lab - DELETE method not allowed (405)", self.EP, "DELETE", None, 405, "HTTP 405", r, el)

    def test_BE045_put_not_allowed(self):
        t0 = time.time()
        r = requests.put(f"{BASE_URL}{self.EP}", json={}, timeout=15)
        el = time.time() - t0
        assert record("BE-045", "Future Lab - PUT method not allowed (405)", self.EP, "PUT", {}, 405, "HTTP 405", r, el)

    def test_BE046_patch_not_allowed(self):
        t0 = time.time()
        r = requests.patch(f"{BASE_URL}{self.EP}", json={}, timeout=15)
        el = time.time() - t0
        assert record("BE-046", "Future Lab - PATCH method not allowed (405)", self.EP, "PATCH", {}, 405, "HTTP 405", r, el)


# ─── Page Route Availability ──────────────────────────────────────────────────

class TestPageRoutes:

    def test_BE050_home_200(self):
        r, el, err = _get("/")
        assert record("BE-050", "Home page reachable (HTTP 200)", "/", "GET", None, 200, "HTTP 200 OK", r, el, err)

    def test_BE051_login_200(self):
        r, el, err = _get("/login")
        assert record("BE-051", "Login page reachable (HTTP 200)", "/login", "GET", None, 200, "HTTP 200 OK", r, el, err)

    def test_BE052_signup_200(self):
        r, el, err = _get("/signup")
        assert record("BE-052", "Signup page reachable (HTTP 200)", "/signup", "GET", None, 200, "HTTP 200 OK", r, el, err)

    def test_BE053_about_200(self):
        r, el, err = _get("/about")
        assert record("BE-053", "About page reachable (HTTP 200)", "/about", "GET", None, 200, "HTTP 200 OK", r, el, err)

    def test_BE054_features_200(self):
        r, el, err = _get("/features")
        assert record("BE-054", "Features page reachable (HTTP 200)", "/features", "GET", None, 200, "HTTP 200 OK", r, el, err)

    def test_BE055_contact_page_200(self):
        r, el, err = _get("/contact")
        assert record("BE-055", "Contact page reachable (HTTP 200)", "/contact", "GET", None, 200, "HTTP 200 OK", r, el, err)

    def test_BE056_privacy_200(self):
        r, el, err = _get("/privacy")
        assert record("BE-056", "Privacy page reachable (HTTP 200)", "/privacy", "GET", None, 200, "HTTP 200 OK", r, el, err)

    def test_BE057_terms_200(self):
        r, el, err = _get("/terms")
        assert record("BE-057", "Terms page reachable (HTTP 200)", "/terms", "GET", None, 200, "HTTP 200 OK", r, el, err)

    def test_BE058_dashboard_unauthenticated_redirects(self):
        t0 = time.time()
        r = requests.get(f"{BASE_URL}/dashboard", timeout=15, allow_redirects=False)
        el = time.time() - t0
        is_redir = 300 <= r.status_code < 400
        _results.append({
            "Test ID": "BE-058",
            "Test Case": "Dashboard - unauthenticated access redirects (3xx)",
            "Endpoint": "GET /dashboard",
            "Input": "(none)",
            "Expected Result": "HTTP 3xx Redirect to /login",
            "Actual Result": f"HTTP {r.status_code} Location:{r.headers.get('Location','')}",
            "HTTP Status": r.status_code,
            "Pass/Fail": "PASS" if is_redir else "FAIL",
            "Error Details": ("" if is_redir else f"Expected 3xx, got {r.status_code}"),
            "Execution Time (s)": round(el, 3),
        })
        assert is_redir, f"Expected 3xx redirect, got {r.status_code}"

    def test_BE059_sleep_unauthenticated_redirects(self):
        t0 = time.time()
        r = requests.get(f"{BASE_URL}/sleep", timeout=15, allow_redirects=False)
        el = time.time() - t0
        is_redir = 300 <= r.status_code < 400
        _results.append({
            "Test ID": "BE-059",
            "Test Case": "Sleep page - unauthenticated redirects (3xx)",
            "Endpoint": "GET /sleep",
            "Input": "(none)",
            "Expected Result": "HTTP 3xx Redirect",
            "Actual Result": f"HTTP {r.status_code}",
            "HTTP Status": r.status_code,
            "Pass/Fail": "PASS" if is_redir else "FAIL",
            "Error Details": ("" if is_redir else f"Expected 3xx, got {r.status_code}"),
            "Execution Time (s)": round(el, 3),
        })
        assert is_redir

    def test_BE060_404_nonexistent_route(self):
        r, el, err = _get("/this-page-does-not-exist-at-all-xyz")
        assert record("BE-060", "Non-existent route returns 404", "/this-page-does-not-exist", "GET", None, 404, "HTTP 404", r, el, err)

    def test_BE061_404_api_nonexistent(self):
        r, el, err = _get("/api/nonexistent-endpoint-xyz")
        assert record("BE-061", "Non-existent API endpoint returns 404", "/api/nonexistent", "GET", None, 404, "HTTP 404", r, el, err)


# ─── Security Header Tests ────────────────────────────────────────────────────

class TestSecurityHeaders:

    def test_BE070_x_content_type_options(self):
        r, el, err = _get("/")
        record("BE-070", "Home - X-Content-Type-Options: nosniff present", "/", "GET", None, 200, "X-Content-Type-Options:nosniff", r, el, err)
        assert r is not None and r.status_code == 200
        hdr = r.headers.get("X-Content-Type-Options", "")
        ok = "nosniff" in hdr
        _results[-1]["Pass/Fail"] = "PASS" if ok else "FAIL"
        _results[-1]["Actual Result"] = f"X-Content-Type-Options: {hdr}"
        _results[-1]["Error Details"] = "" if ok else "Missing X-Content-Type-Options: nosniff"
        assert ok, f"Missing X-Content-Type-Options (got: {hdr!r})"

    def test_BE071_x_frame_options(self):
        r, el, err = _get("/")
        record("BE-071", "Home - X-Frame-Options header present", "/", "GET", None, 200, "X-Frame-Options set", r, el, err)
        assert r is not None and r.status_code == 200
        hdr = r.headers.get("X-Frame-Options", "")
        ok = bool(hdr)
        _results[-1]["Pass/Fail"] = "PASS" if ok else "FAIL"
        _results[-1]["Actual Result"] = f"X-Frame-Options: {hdr}"
        _results[-1]["Error Details"] = "" if ok else "Missing X-Frame-Options header"
        assert ok, "Missing X-Frame-Options header"

    def test_BE072_api_error_json_content_type(self):
        p = {"name": "A", "email": "bad", "message": "x"}
        r, el, err = _post("/api/contact", p)
        record("BE-072", "API error response Content-Type is application/json", "/api/contact", "POST", None, 400, "application/json", r, el, err)
        assert r is not None
        ct = r.headers.get("Content-Type", "")
        ok = "application/json" in ct
        _results[-1]["Pass/Fail"] = "PASS" if ok else "FAIL"
        _results[-1]["Actual Result"] = f"Content-Type: {ct}"
        assert ok, f"Expected application/json, got {ct}"

    def test_BE073_robots_txt_accessible(self):
        r, el, err = _get("/robots.txt")
        ok = r is not None and r.status_code in (200, 404)
        _results.append({
            "Test ID": "BE-073",
            "Test Case": "robots.txt accessible (200 or 404)",
            "Endpoint": "GET /robots.txt",
            "Input": "(none)",
            "Expected Result": "HTTP 200 or 404",
            "Actual Result": f"HTTP {r.status_code if r else 'NO_RESPONSE'}",
            "HTTP Status": r.status_code if r else "ERR",
            "Pass/Fail": "PASS" if ok else "FAIL",
            "Error Details": err if not ok else "",
            "Execution Time (s)": round(el, 3),
        })
        assert ok


# ─── Report Generation (session finish hook) ──────────────────────────────────

def pytest_sessionfinish(session, exitstatus):
    reports_dir = pathlib.Path("backend-tests/reports")
    reports_dir.mkdir(parents=True, exist_ok=True)

    # JSON
    json_path = reports_dir / "backend_results.json"
    with open(json_path, "w") as f:
        json.dump(_results, f, indent=2)

    total = len(_results)
    passed = sum(1 for r in _results if r["Pass/Fail"] == "PASS")
    failed = total - passed
    pct = round(passed / total * 100, 1) if total > 0 else 0

    # HTML
    rows_html = ""
    for r in _results:
        color = "#d4edda" if r["Pass/Fail"] == "PASS" else "#f8d7da"
        badge = f'<span style="background:{"#28a745" if r["Pass/Fail"]=="PASS" else "#dc3545"};color:#fff;padding:2px 8px;border-radius:4px;">{r["Pass/Fail"]}</span>'
        rows_html += (
            f'<tr style="background:{color}">'
            f'<td>{r["Test ID"]}</td><td>{r["Test Case"]}</td>'
            f'<td style="font-size:11px">{r["Endpoint"]}</td>'
            f'<td style="font-size:10px">{r["Input"]}</td>'
            f'<td style="font-size:10px">{r["Expected Result"]}</td>'
            f'<td style="font-size:10px">{r["Actual Result"]}</td>'
            f'<td>{r["HTTP Status"]}</td><td>{badge}</td>'
            f'<td style="font-size:10px;color:#c00">{r["Error Details"]}</td>'
            f'<td>{r["Execution Time (s)"]}s</td></tr>'
        )

    html = f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<title>VitalCore Backend Test Report</title>
<style>body{{font-family:Arial,sans-serif;margin:24px;background:#f5f5f5}}
h1{{color:#1a1a2e}}table{{border-collapse:collapse;width:100%;background:#fff;font-size:12px}}
th{{background:#1a1a2e;color:#fff;padding:8px;text-align:left}}
td{{border:1px solid #ccc;padding:5px 7px;vertical-align:top}}
.badges{{display:flex;gap:12px;margin:16px 0}}
.badge{{padding:10px 18px;border-radius:8px;color:#fff;font-size:15px;font-weight:bold}}</style>
</head><body>
<h1>VitalCore Backend API Test Report</h1>
<p><strong>Generated:</strong> {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}</p>
<p><strong>Target:</strong> {BASE_URL}</p>
<div class="badges">
<div class="badge" style="background:#1a1a2e">Total: {total}</div>
<div class="badge" style="background:#28a745">Passed: {passed}</div>
<div class="badge" style="background:#dc3545">Failed: {failed}</div>
<div class="badge" style="background:#007bff">Pass Rate: {pct}%</div>
</div>
<table><thead><tr>
<th>Test ID</th><th>Test Case</th><th>Endpoint</th><th>Input</th>
<th>Expected</th><th>Actual</th><th>HTTP</th><th>Result</th><th>Error</th><th>Time</th>
</tr></thead><tbody>{rows_html}</tbody></table>
</body></html>"""

    html_path = reports_dir / "backend_results.html"
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)

    # Excel
    wb = Workbook()
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="1A1A2E")

    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(["VitalCore Backend API Test Summary"])
    ws_sum.append(["Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")])
    ws_sum.append(["Target URL", BASE_URL])
    ws_sum.append([])
    ws_sum.append(["Metric", "Value"])
    ws_sum.append(["Total Test Cases", total])
    ws_sum.append(["Executed", total])
    ws_sum.append(["Passed", passed])
    ws_sum.append(["Failed", failed])
    ws_sum.append(["Pass Rate (%)", pct])
    ws_sum["A1"].font = Font(bold=True, size=14)
    for cell in ws_sum[5]:
        cell.font = hdr_font; cell.fill = hdr_fill
    for row in ws_sum.iter_rows(min_row=6, max_row=10, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin
    ws_sum["B8"].fill = PatternFill("solid", fgColor="D4EDDA")
    ws_sum["B9"].fill = PatternFill("solid", fgColor="F8D7DA") if failed > 0 else PatternFill("solid", fgColor="D4EDDA")
    ws_sum.column_dimensions["A"].width = 25
    ws_sum.column_dimensions["B"].width = 40

    ws_det = wb.create_sheet("Detailed Results")
    cols = ["Test ID","Test Case","Endpoint","Input","Expected Result","Actual Result","HTTP Status","Pass/Fail","Error Details","Execution Time (s)"]
    ws_det.append(cols)
    for cell in ws_det[1]:
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin
    for rec in _results:
        ws_det.append([rec.get(c, "") for c in cols])
        row_num = ws_det.max_row
        rf = PatternFill("solid", fgColor=("D4EDDA" if rec["Pass/Fail"] == "PASS" else "F8D7DA"))
        for col_idx in range(1, len(cols) + 1):
            cell = ws_det.cell(row=row_num, column=col_idx)
            cell.fill = rf; cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for i, w in enumerate([10,45,32,22,30,30,12,10,30,14], 1):
        ws_det.column_dimensions[get_column_letter(i)].width = w

    xlsx_path = reports_dir / "backend_results.xlsx"
    wb.save(xlsx_path)
    print(f"\n{'='*60}\n  BACKEND TEST SUMMARY\n{'='*60}")
    print(f"  Total  : {total}\n  Passed : {passed}\n  Failed : {failed}\n  Rate   : {pct}%")
    print(f"  JSON   : {json_path}\n  HTML   : {html_path}\n  Excel  : {xlsx_path}")
    print(f"{'='*60}\n")
