import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from .config import EXCEL_REPORT_PATH

def generate_excel_report(test_results, total_duration_seconds, start_time_str=None, end_time_str=None):
    """
    Generates a professional 3-sheet Excel report matching prompt specifications.
    Sheets:
    1. Test Results
    2. Summary
    3. Module Summary
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Styles
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Light Green
    pass_font = Font(name="Segoe UI", size=10, bold=True, color="166534")
    
    fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Light Red
    fail_font = Font(name="Segoe UI", size=10, bold=True, color="991B1B")

    skip_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Light Yellow
    skip_font = Font(name="Segoe UI", size=10, bold=True, color="92400E")

    regular_font = Font(name="Segoe UI", size=10)
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # Calculate Totals
    total_tests = len(test_results)
    passed_tests = sum(1 for item in test_results if item.get('status') == 'PASS')
    failed_tests = sum(1 for item in test_results if item.get('status') == 'FAIL')
    skipped_tests = sum(1 for item in test_results if item.get('status') == 'SKIPPED')
    
    pass_pct = round((passed_tests / total_tests * 100), 2) if total_tests > 0 else 0.0
    fail_pct = round((failed_tests / total_tests * 100), 2) if total_tests > 0 else 0.0

    if not start_time_str:
        start_time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if not end_time_str:
        end_time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # -------------------------------------------------------------
    # SHEET 1: Test Results
    # -------------------------------------------------------------
    ws1 = wb.create_sheet(title="Test Results")
    headers1 = [
        "Test ID",
        "Module",
        "Test Case",
        "Preconditions",
        "Steps",
        "Expected Result",
        "Actual Result",
        "Status",
        "Error/Exception",
        "URL",
        "Execution Time (s)",
        "Screenshot"
    ]
    ws1.append(headers1)

    for col_idx, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for item in test_results:
        status = item.get("status", "PASS")
        row = [
            item.get("test_id", ""),
            item.get("module", ""),
            item.get("name", ""),
            item.get("preconditions", ""),
            item.get("steps", ""),
            item.get("expected_result", item.get("description", "")),
            item.get("actual_result", ""),
            status,
            item.get("error", ""),
            item.get("url", ""),
            round(item.get("duration", 0), 2),
            item.get("screenshot", ""),
        ]
        ws1.append(row)
        current_row = ws1.max_row
        
        # Format Status Cell (column 8)
        status_cell = ws1.cell(row=current_row, column=8)
        if status == "PASS":
            status_cell.fill = pass_fill
            status_cell.font = pass_font
        elif status == "FAIL":
            status_cell.fill = fail_fill
            status_cell.font = fail_font
        else:
            status_cell.fill = skip_fill
            status_cell.font = skip_font
        status_cell.alignment = Alignment(horizontal="center")

        for col in range(1, len(headers1) + 1):
            c = ws1.cell(row=current_row, column=col)
            c.border = thin_border
            if col != 8:
                c.font = regular_font

    # -------------------------------------------------------------
    # SHEET 2: Summary
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Summary")
    ws2.append(["VitalCore Selenium E2E Test Execution Summary"])
    ws2.merge_cells("A1:B1")
    ws2.cell(row=1, column=1).font = Font(name="Segoe UI", size=14, bold=True, color="1E293B")
    ws2.append([])

    summary_rows = [
        ("Total Tests", total_tests),
        ("Passed", passed_tests),
        ("Failed", failed_tests),
        ("Skipped", skipped_tests),
        ("Pass Percentage", f"{pass_pct}%"),
        ("Fail Percentage", f"{fail_pct}%"),
        ("Execution Start Time", start_time_str),
        ("Execution End Time", end_time_str),
        ("Total Execution Duration", f"{round(total_duration_seconds, 2)}s"),
    ]

    for label, val in summary_rows:
        ws2.append([label, val])
        r = ws2.max_row
        c1 = ws2.cell(row=r, column=1)
        c2 = ws2.cell(row=r, column=2)
        c1.font = bold_font
        c2.font = regular_font
        c1.border = thin_border
        c2.border = thin_border

    # -------------------------------------------------------------
    # SHEET 3: Module Summary
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="Module Summary")
    headers3 = ["Module", "Total", "Passed", "Failed", "Pass %"]
    ws3.append(headers3)
    for col_idx, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font

    modules_map = {}
    for item in test_results:
        m = item.get("module", "General")
        if m not in modules_map:
            modules_map[m] = {"total": 0, "pass": 0, "fail": 0}
        modules_map[m]["total"] += 1
        st = item.get("status", "PASS")
        if st == "PASS":
            modules_map[m]["pass"] += 1
        elif st == "FAIL":
            modules_map[m]["fail"] += 1

    for mod_name, stats in modules_map.items():
        m_pass_pct = round((stats["pass"] / stats["total"] * 100), 2) if stats["total"] > 0 else 0.0
        ws3.append([mod_name, stats["total"], stats["pass"], stats["fail"], f"{m_pass_pct}%"])
        r = ws3.max_row
        for col in range(1, 6):
            c = ws3.cell(row=r, column=col)
            c.border = thin_border
            c.font = regular_font

    # Auto-adjust column widths across all sheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(EXCEL_REPORT_PATH)
    return EXCEL_REPORT_PATH
