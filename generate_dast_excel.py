import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

def generate_excel():
    with open('report.json', 'r') as f:
        data = json.load(f)
        
    wb = openpyxl.Workbook()
    
    # Fonts and Fills
    title_font = Font(bold=True, size=16, color="FFFFFF")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    bold_font = Font(bold=True)
    
    bg_dark = PatternFill("solid", fgColor="1E293B")
    bg_black = PatternFill("solid", fgColor="0F172A")
    bg_pass = PatternFill("solid", fgColor="10B981")
    bg_fail = PatternFill("solid", fgColor="EF4444")
    
    thin_border = Border(left=Side(style='thin', color='CCCCCC'),
                         right=Side(style='thin', color='CCCCCC'),
                         top=Side(style='thin', color='CCCCCC'),
                         bottom=Side(style='thin', color='CCCCCC'))
    
    # -----------------------------
    # Sheet 1: Executive Dashboard
    # -----------------------------
    ws1 = wb.active
    ws1.title = "Executive Dashboard"
    
    ws1.merge_cells("B2:G2")
    ws1['B2'] = "VITALCORE AI - DAST SECURITY REPORT"
    ws1['B2'].font = title_font
    ws1['B2'].fill = bg_black
    ws1['B2'].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 30
    
    ws1['B3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} UTC"
    ws1['B4'] = "Target: http://localhost:3000"
    
    total = len(data)
    failed = len([d for d in data if d.get('finding')])
    passed = total - failed
    pass_rate = f"{(passed/total)*100:.1f}%" if total > 0 else "0%"
    
    ws1['B6'] = "TOTAL TESTS"
    ws1['C6'] = "PASSED"
    ws1['D6'] = "FAILED"
    ws1['E6'] = "PASS RATE"
    for col in ['B', 'C', 'D', 'E']:
        ws1[f'{col}6'].font = header_font
        ws1[f'{col}6'].fill = bg_dark
        ws1[f'{col}6'].alignment = Alignment(horizontal="center")
        
    ws1['B7'] = total
    ws1['C7'] = passed
    ws1['D7'] = failed
    ws1['E7'] = pass_rate
    for col in ['B', 'C', 'D', 'E']:
        ws1[f'{col}7'].font = bold_font
        ws1[f'{col}7'].alignment = Alignment(horizontal="center")
        ws1[f'{col}7'].border = thin_border
        
    # Severity
    ws1['B10'] = "FINDINGS BY SEVERITY"
    ws1['B10'].font = bold_font
    
    sevs = ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW', 'INFO']
    for idx, sev in enumerate(sevs):
        ws1.cell(row=11, column=2+idx).value = sev
        ws1.cell(row=11, column=2+idx).font = header_font
        ws1.cell(row=11, column=2+idx).fill = bg_dark
        ws1.cell(row=11, column=2+idx).alignment = Alignment(horizontal="center")
        
        count = len([d for d in data if d.get('finding') and str(d.get('severity', '')).upper() == sev])
        ws1.cell(row=12, column=2+idx).value = count
        ws1.cell(row=12, column=2+idx).alignment = Alignment(horizontal="center")
        ws1.cell(row=12, column=2+idx).border = thin_border

    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 15
    ws1.column_dimensions['F'].width = 15
    ws1.column_dimensions['G'].width = 15

    # -----------------------------
    # Sheet 2: All Test Results
    # -----------------------------
    ws2 = wb.create_sheet("All Test Results")
    headers = ['#', 'Category', 'Test ID', 'Test Name', 'URL', 'Method', 'Status Code', 'Severity', 'Result', 'Detail', 'Response Snippet', 'Timestamp']
    for col_num, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = bg_dark
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    for idx, row_data in enumerate(data, 1):
        is_fail = row_data.get('finding')
        result = "FAIL" if is_fail else "PASS"
        
        vals = [
            idx,
            row_data.get('test_category', 'Unknown'),
            f"T{idx:02d}",
            f"{row_data.get('test_category', '')} check",
            row_data.get('endpoint', ''),
            row_data.get('method', ''),
            row_data.get('status', ''),
            row_data.get('severity', 'INFO') if is_fail else 'INFO',
            result,
            row_data.get('note', ''),
            '',
            row_data.get('timestamp', '')
        ]
        
        for col_num, val in enumerate(vals, 1):
            cell = ws2.cell(row=idx+1, column=col_num)
            cell.value = val
            cell.border = thin_border
            if col_num == 9: # Result column
                cell.fill = bg_fail if is_fail else bg_pass
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")

    for col_letter, width in zip(['A','B','C','D','E','F','G','H','I','J','K','L'], [5, 20, 10, 25, 30, 10, 15, 15, 15, 40, 20, 25]):
        ws2.column_dimensions[col_letter].width = width

    # -----------------------------
    # Sheet 3: Failures & Findings
    # -----------------------------
    ws3 = wb.create_sheet("Failures & Findings")
    for col_num, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = bg_dark
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    fail_idx = 1
    for row_data in data:
        if row_data.get('finding'):
            vals = [
                fail_idx,
                row_data.get('test_category', 'Unknown'),
                f"F{fail_idx:02d}",
                f"{row_data.get('test_category', '')} check",
                row_data.get('endpoint', ''),
                row_data.get('method', ''),
                row_data.get('status', ''),
                row_data.get('severity', 'HIGH'),
                "FAIL",
                row_data.get('note', ''),
                '',
                row_data.get('timestamp', '')
            ]
            for col_num, val in enumerate(vals, 1):
                cell = ws3.cell(row=fail_idx+1, column=col_num)
                cell.value = val
                cell.border = thin_border
                if col_num == 9:
                    cell.fill = bg_fail
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center")
            fail_idx += 1

    for col_letter, width in zip(['A','B','C','D','E','F','G','H','I','J','K','L'], [5, 20, 10, 25, 30, 10, 15, 15, 15, 40, 20, 25]):
        ws3.column_dimensions[col_letter].width = width

    # -----------------------------
    # Sheet 4: Category Summary
    # -----------------------------
    ws4 = wb.create_sheet("Category Summary")
    cat_headers = ["Category", "Total Tests", "Passed", "Failed", "Pass Rate"]
    for col_num, header in enumerate(cat_headers, 1):
        cell = ws4.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = bg_dark
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    categories = {}
    for d in data:
        cat = d.get('test_category', 'Unknown')
        if cat not in categories:
            categories[cat] = {'total': 0, 'pass': 0, 'fail': 0}
        categories[cat]['total'] += 1
        if d.get('finding'):
            categories[cat]['fail'] += 1
        else:
            categories[cat]['pass'] += 1
            
    for idx, (cat, stats) in enumerate(categories.items(), 2):
        pr = f"{(stats['pass']/stats['total'])*100:.1f}%"
        vals = [cat, stats['total'], stats['pass'], stats['fail'], pr]
        for col_num, val in enumerate(vals, 1):
            cell = ws4.cell(row=idx, column=col_num)
            cell.value = val
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_num > 1 else "left")

    for col_letter, width in zip(['A','B','C','D','E'], [25, 15, 15, 15, 15]):
        ws4.column_dimensions[col_letter].width = width

    # -----------------------------
    # Sheet 5: Remediation Guide
    # -----------------------------
    ws5 = wb.create_sheet("Remediation Guide")
    ws5.column_dimensions['A'].width = 25
    ws5.column_dimensions['B'].width = 80
    
    ws5['A1'] = "Category"
    ws5['B1'] = "Remediation Advice"
    ws5['A1'].font = header_font
    ws5['B1'].font = header_font
    ws5['A1'].fill = bg_dark
    ws5['B1'].fill = bg_dark
    
    advice = {
        "AuthZ / RBAC": "Ensure all API endpoints validate the Supabase authentication token instead of returning 401 globally, or fix middleware pass-through logic.",
        "AuthN Bypass": "Maintain strict 401 blocks on requests without valid Authorization headers.",
        "Injection Probe": "Continue using parameterized queries and ORMs to prevent SQL/NoSQL injection.",
        "Rate Limiting": "Ensure rate limiting is globally active to prevent DDoS and brute-force attacks.",
        "Token Tampering": "Always securely verify JWT signatures using the proper secret key."
    }
    
    for idx, (cat, text) in enumerate(advice.items(), 2):
        ws5.cell(row=idx, column=1, value=cat).border = thin_border
        ws5.cell(row=idx, column=2, value=text).border = thin_border
        ws5.cell(row=idx, column=2).alignment = Alignment(wrap_text=True)

    timestamp = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    report_name = f"DAST_Report_VitalCore_{timestamp}.xlsx"
    wb.save(report_name)
    print(f"Successfully generated {report_name}")

if __name__ == "__main__":
    generate_excel()
