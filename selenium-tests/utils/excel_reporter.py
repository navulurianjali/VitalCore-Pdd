import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from .config import EXCEL_REPORT_PATH

def generate_excel_report(test_results, total_duration_seconds):
    """
    Generates a professional 5-sheet Excel report using openpyxl.
    Sheets:
    1. Test Results
    2. Summary
    3. Module Summary
    4. Failed Tests
    5. Execution Details
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Styles
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Light Green
    pass_font = Font(name="Segoe UI", size=10, bold=True, color="166534")
    
    fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Light Red
    fail_font = Font(name="Segoe UI", size=10, bold=True, color="991B1B")

    skip_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Light Yellow
    skip_font = Font(name="Segoe UI", size=10, bold=True, color="92400E")

    regular_font = Font(name="Segoe UI", size=10)
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # Calculate Totals
    total_tests = len(test_results)
    passed_tests = sum(1 for item in test_results if item['status'] == 'PASS')
    failed_tests = sum(1 for item in test_results if item['status'] == 'FAIL')
    skipped_tests = sum(1 for item in test_results if item['status'] == 'SKIPPED')
    
    pass_pct = round((passed_tests / total_tests * 100), 2) if total_tests > 0 else 0.0
    fail_pct = round((failed_tests / total_tests * 100), 2) if total_tests > 0 else 0.0

    # -------------------------------------------------------------
    # SHEET 1: Test Results
    # -------------------------------------------------------------
    ws1 = wb.create_sheet(title="Test Results")
    headers1 = ["Test ID", "Module", "Test Case", "Description", "Status", "Execution Time (s)", "Screenshot", "Error Message"]
    ws1.append(headers1)

    for col_idx, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for item in test_results:
        row = [
            item.get("test_id", ""),
            item.get("module", ""),
            item.get("name", ""),
            item.get("description", ""),
            item.get("status", "PASS"),
            round(item.get("duration", 0), 2),
            item.get("screenshot", ""),
            item.get("error", ""),
        ]
        ws1.append(row)
        current_row = ws1.max_row
        
        # Format Status Cell
        status_cell = ws1.cell(row=current_row, column=5)
        if item.get("status") == "PASS":
            status_cell.fill = pass_fill
            status_cell.font = pass_font
        elif item.get("status") == "FAIL":
            status_cell.fill = fail_fill
            status_cell.font = fail_font
        else:
            status_cell.fill = skip_fill
            status_cell.font = skip_font
        status_cell.alignment = Alignment(horizontal="center")

        for col in range(1, len(headers1) + 1):
            c = ws1.cell(row=current_row, column=col)
            c.border = thin_border
            if col != 5:
                c.font = regular_font

    # -------------------------------------------------------------
    # SHEET 2: Summary
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Summary")
    ws2.append(["Selenium Test Suite Execution Summary"])
    ws2.merge_cells("A1:B1")
    ws2.cell(row=1, column=1).font = Font(name="Segoe UI", size=14, bold=True, color="1E293B")
    ws2.append([])

    summary_rows = [
        ("Total Test Cases", total_tests),
        ("Passed Tests", passed_tests),
        ("Failed Tests", failed_tests),
        ("Skipped Tests", skipped_tests),
        ("Pass Percentage", f"{pass_pct}%"),
        ("Fail Percentage", f"{fail_pct}%"),
        ("Execution Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Total Duration (seconds)", round(total_duration_seconds, 2)),
    ]

    for label, val in summary_rows:
        ws2.append([label, val])
        r = ws2.max_row
        c1 = ws2.cell(row=r, column=1)
        c2 = ws2.cell(row=r, column=2)
        c1.font = bold_font
        c2.font = regular_font
        c1.border = thin_border
        c2.border = thin_border

    # -------------------------------------------------------------
    # SHEET 3: Module Summary
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="Module Summary")
    headers3 = ["Module Name", "Total Tests", "Passed", "Failed", "Skipped", "Pass %"]
    ws3.append(headers3)
    for col_idx, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font

    modules_map = {}
    for item in test_results:
        m = item.get("module", "General")
        if m not in modules_map:
            modules_map[m] = {"total": 0, "pass": 0, "fail": 0, "skip": 0}
        modules_map[m]["total"] += 1
        st = item.get("status", "PASS")
        if st == "PASS":
            modules_map[m]["pass"] += 1
        elif st == "FAIL":
            modules_map[m]["fail"] += 1
        else:
            modules_map[m]["skip"] += 1

    for mod_name, stats in modules_map.items():
        m_pass_pct = round((stats["pass"] / stats["total"] * 100), 2) if stats["total"] > 0 else 0.0
        ws3.append([mod_name, stats["total"], stats["pass"], stats["fail"], stats["skip"], f"{m_pass_pct}%"])
        r = ws3.max_row
        for col in range(1, 7):
            c = ws3.cell(row=r, column=col)
            c.border = thin_border
            c.font = regular_font

    # -------------------------------------------------------------
    # SHEET 4: Failed Tests
    # -------------------------------------------------------------
    ws4 = wb.create_sheet(title="Failed Tests")
    headers4 = ["Test ID", "Module", "Test Case", "Error Message", "Screenshot Path"]
    ws4.append(headers4)
    for col_idx, h in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col_idx)
        cell.fill = PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid")
        cell.font = header_font

    for item in test_results:
        if item.get("status") == "FAIL":
            ws4.append([
                item.get("test_id", ""),
                item.get("module", ""),
                item.get("name", ""),
                item.get("error", ""),
                item.get("screenshot", ""),
            ])
            r = ws4.max_row
            for col in range(1, 6):
                c = ws4.cell(row=r, column=col)
                c.border = thin_border
                c.font = regular_font

    # -------------------------------------------------------------
    # SHEET 5: Execution Details
    # -------------------------------------------------------------
    ws5 = wb.create_sheet(title="Execution Details")
    headers5 = ["Property", "Value"]
    ws5.append(headers5)
    for col_idx, h in enumerate(headers5, 1):
        cell = ws5.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font

    exec_details = [
        ("Framework", "Python + Selenium + Pytest"),
        ("Execution Environment", "Local / CI"),
        ("Browser Engine", "Chrome Headless / Desktop"),
        ("Report Generated At", datetime.now().isoformat()),
        ("Target Test Cases", "330"),
    ]
    for p, v in exec_details:
        ws5.append([p, v])
        r = ws5.max_row
        ws5.cell(row=r, column=1).font = bold_font
        ws5.cell(row=r, column=2).font = regular_font
        ws5.cell(row=r, column=1).border = thin_border
        ws5.cell(row=r, column=2).border = thin_border

    # Auto-adjust column widths across all sheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(EXCEL_REPORT_PATH)
    return EXCEL_REPORT_PATH
